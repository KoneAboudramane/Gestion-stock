import io
from decimal import Decimal

from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from catalogue.models import Categorie, Produit, Variante
from clients.models import Client
from comptes.models import Role, Utilisateur
from comptes.services import inscrire_boutique
from stock.models import Depot, MouvementStock
from stock.services import appliquer_mouvement
from ventes.services import creer_vente, annuler_vente


class RapportsTestsBase(APITestCase):
    def setUp(self):
        self.boutique, self.patron = inscrire_boutique(
            {"nom": "Boutique A"}, {"username": "patronA", "password": "UnMotDePasseSolide123"}
        )
        self.depot = Depot.objects.create(boutique=self.boutique, nom="Magasin")
        self.categorie = Categorie.objects.create(boutique=self.boutique, nom="Alimentation")

        produit1 = Produit.objects.create(boutique=self.boutique, nom="Riz", categorie=self.categorie)
        self.variante1 = Variante.objects.create(produit=produit1, prix_achat=8000, prix_vente=12000, seuil_alerte=5)
        produit2 = Produit.objects.create(boutique=self.boutique, nom="Savon")
        self.variante2 = Variante.objects.create(produit=produit2, prix_achat=200, prix_vente=350, seuil_alerte=10)

        appliquer_mouvement(self.variante1, self.depot, MouvementStock.Type.ENTREE, 20)
        appliquer_mouvement(self.variante2, self.depot, MouvementStock.Type.ENTREE, 3)  # <= seuil -> rupture

        self.vente1 = creer_vente(
            boutique=self.boutique, depot=self.depot, utilisateur=self.patron, client=None, statut="payee",
            lignes_donnees=[{"variante": self.variante1, "quantite": 2}],
            paiements_donnees=[{"mode": "especes", "montant": 24000}],
        )
        self.vente2 = creer_vente(
            boutique=self.boutique, depot=self.depot, utilisateur=self.patron, client=None, statut="payee",
            lignes_donnees=[{"variante": self.variante2, "quantite": 3}],
            paiements_donnees=[{"mode": "mobile_money", "montant": 1050}],
        )
        # Vente annulée : ne doit compter dans aucun rapport.
        self.vente_annulee = creer_vente(
            boutique=self.boutique, depot=self.depot, utilisateur=self.patron, client=None, statut="payee",
            lignes_donnees=[{"variante": self.variante1, "quantite": 1}],
            paiements_donnees=[{"mode": "especes", "montant": 12000}],
        )
        annuler_vente(self.vente_annulee, utilisateur=self.patron)

        self.client.force_authenticate(user=self.patron)


class SyntheseEtBeneficeTests(RapportsTestsBase):
    def test_synthese_exclut_les_ventes_annulees(self):
        reponse = self.client.get(reverse("rapport-synthese-ventes"), {"periode": "jour"})
        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        donnees = reponse.data[0]
        # vente1: 24000 (cout 16000 -> benefice 8000) + vente2: 1050 (cout 600 -> benefice 450)
        self.assertEqual(donnees["total_net"], Decimal("25050"))
        self.assertEqual(donnees["nombre_ventes"], 2)
        self.assertEqual(donnees["benefice_total"], Decimal("8450"))


class TopProduitsTests(RapportsTestsBase):
    def test_top_produits_ordre_desc_et_asc(self):
        reponse = self.client.get(reverse("rapport-top-produits"), {"periode": "jour", "ordre": "desc"})
        self.assertEqual(reponse.data[0]["produit"], "Savon")  # 3 unités vendues > 2

        reponse_asc = self.client.get(reverse("rapport-top-produits"), {"periode": "jour", "ordre": "asc"})
        self.assertEqual(reponse_asc.data[0]["produit"], "Riz")


class ValeurStockTests(RapportsTestsBase):
    def test_valeur_stock_et_ruptures(self):
        reponse = self.client.get(reverse("rapport-valeur-stock"))
        donnees = reponse.data[0]
        # riz: 20 - 2 (vente1) - 1 (vente annulee) + 1 (restauration a l'annulation) = 18 * 8000 = 144000
        # savon: 3 - 3 (vente2) = 0 * 200 = 0
        self.assertEqual(donnees["valeur_achat"], Decimal("144000"))
        self.assertEqual(donnees["nombre_ruptures"], 1)


class VentesParDimensionTests(RapportsTestsBase):
    def test_par_vendeur_categorie_mode_paiement(self):
        reponse = self.client.get(reverse("rapport-ventes-par-vendeur"), {"periode": "jour"})
        self.assertEqual(reponse.data[0]["utilisateur"], "patronA")
        self.assertEqual(reponse.data[0]["nombre_ventes"], 2)

        reponse = self.client.get(reverse("rapport-ventes-par-categorie"), {"periode": "jour"})
        categories = {ligne["categorie"] for ligne in reponse.data}
        self.assertIn("Alimentation", categories)
        self.assertIn("Sans catégorie", categories)

        reponse = self.client.get(reverse("rapport-ventes-par-mode-paiement"), {"periode": "jour"})
        modes = {ligne["mode"]: ligne["total"] for ligne in reponse.data}
        self.assertEqual(modes["especes"], Decimal("24000"))
        self.assertEqual(modes["mobile_money"], Decimal("1050"))


class VentesParJourTests(RapportsTestsBase):
    def test_ventes_par_jour_exclut_les_ventes_annulees(self):
        reponse = self.client.get(reverse("rapport-ventes-par-jour"), {"periode": "jour"})
        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        self.assertEqual(len(reponse.data), 1)
        self.assertEqual(reponse.data[0]["total_net"], Decimal("25050"))


class TopClientsTests(RapportsTestsBase):
    def test_top_clients_exclut_les_ventes_sans_client(self):
        client_fidele = Client.objects.create(boutique=self.boutique, nom="Awa Traoré")
        creer_vente(
            boutique=self.boutique, depot=self.depot, utilisateur=self.patron, client=client_fidele, statut="payee",
            lignes_donnees=[{"variante": self.variante1, "quantite": 1}],
            paiements_donnees=[{"mode": "especes", "montant": 12000}],
        )

        reponse = self.client.get(reverse("rapport-top-clients"), {"periode": "jour"})
        self.assertEqual(reponse.status_code, status.HTTP_200_OK, reponse.data)
        # vente1 et vente2 (setUp) n'ont pas de client -> seule la vente de Awa apparaît.
        self.assertEqual(len(reponse.data), 1)
        self.assertEqual(reponse.data[0]["client_nom"], "Awa Traoré")
        self.assertEqual(reponse.data[0]["total_net"], Decimal("12000"))


class ExportTests(RapportsTestsBase):
    def test_export_csv(self):
        reponse = self.client.get(reverse("rapport-top-produits"), {"export": "csv"})
        self.assertEqual(reponse.status_code, status.HTTP_200_OK)
        self.assertEqual(reponse["Content-Type"], "text/csv")
        contenu = reponse.content.decode("utf-8")
        self.assertEqual(len(contenu.strip().splitlines()), 3)  # entête + 2 produits

    def test_export_xlsx(self):
        reponse = self.client.get(reverse("rapport-top-produits"), {"export": "xlsx"})
        self.assertEqual(
            reponse["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        classeur = load_workbook(io.BytesIO(reponse.content))
        feuille = classeur.active
        self.assertEqual(feuille.max_row, 3)

    def test_export_pdf(self):
        reponse = self.client.get(reverse("rapport-top-produits"), {"export": "pdf"})
        self.assertEqual(reponse["Content-Type"], "application/pdf")
        self.assertTrue(reponse.content.startswith(b"%PDF"))


class PermissionsRapportsTests(RapportsTestsBase):
    def test_caissier_refuse_sur_tous_les_rapports(self):
        caissier = Utilisateur(
            boutique=self.boutique,
            role=Role.objects.get(boutique=self.boutique, nom="Caissier"),
            username="caissierA",
        )
        caissier.set_password("UnMotDePasseSolide123")
        caissier.save()
        self.client.force_authenticate(user=caissier)

        for nom_url in (
            "rapport-synthese-ventes", "rapport-top-produits", "rapport-valeur-stock",
            "rapport-ventes-par-vendeur", "rapport-ventes-par-categorie", "rapport-ventes-par-mode-paiement",
            "rapport-ventes-par-jour", "rapport-top-clients",
        ):
            reponse = self.client.get(reverse(nom_url))
            self.assertEqual(reponse.status_code, status.HTTP_403_FORBIDDEN, nom_url)
